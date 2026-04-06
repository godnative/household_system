import sys
from PyQt5.QtWidgets import *
from PyQt5.QtWebEngineWidgets import QWebEngineView  # 浏览器引擎
from qfluentwidgets import PushButton
from openpyxl import load_workbook

class ExcelWebEngineViewer(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("XLSX → Web引擎渲染富文本表格")
        self.resize(1000, 700)

        # 主容器
        cw = QWidget()
        self.setCentralWidget(cw)
        layout = QVBoxLayout(cw)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        # 按钮
        self.btn_open = PushButton("📄 打开 XLSX")
        self.btn_import = PushButton("📂 导入 HTML")
        self.btn_export = PushButton("💾 导出 HTML")

        self.btn_open.clicked.connect(self.open_xlsx)
        self.btn_import.clicked.connect(self.import_html)
        self.btn_export.clicked.connect(self.export_html)

        # 按钮布局
        btn_layout = QHBoxLayout()
        btn_layout.addWidget(self.btn_open)
        btn_layout.addWidget(self.btn_import)
        btn_layout.addWidget(self.btn_export)
        layout.addLayout(btn_layout)

        # ========================
        # 核心：Web 引擎浏览器控件
        # ========================
        self.web = QWebEngineView()
        layout.addWidget(self.web)

    # ------------------------------
    # 打开 XLSX → 生成 HTML → 浏览器渲染
    # ------------------------------
    def open_xlsx(self):
        path, _ = QFileDialog.getOpenFileName(filter="Excel (*.xlsx)")
        if not path:
            return
        html = self.excel_to_html(path)
        self.web.setHtml(html)  # 浏览器引擎渲染

    # ------------------------------
    # Excel 转 HTML（合并单元格 + 宽度30px）
    # ------------------------------
    def excel_to_html(self, file_path):
        wb = load_workbook(file_path)
        sheet = wb.active

        # 读取合并单元格
        merged_ranges = sheet.merged_cells.ranges
        merged_set = set()
        merge_info = {}

        for m in merged_ranges:
            mr, mc, Mr, Mc = m.min_row, m.min_col, m.max_row, m.max_col
            rspan = Mr - mr + 1
            cspan = Mc - mc + 1
            merge_info[(mr, mc)] = (rspan, cspan)
            for r in range(mr, Mr+1):
                for c in range(mc, Mc+1):
                    if r != mr or c != mc:
                        merged_set.add((r, c))

        # 完整 HTML（浏览器标准格式）
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-size:14px; padding:20px; }}
                table {{
                    border-collapse:collapse;
                    width:100%;
                }}
                th, td {{
                    border:1px solid #999;
                    padding:8px;
                    width:30px;  /* 固定宽度 30px */
                    text-align:left;
                }}
                th {{
                    background:#f2f2f2;
                    font-weight:bold;
                }}
            </style>
        </head>
        <body>
            <h3>Web引擎渲染表格</h3>
            <table>
        """

        for row_idx, row in enumerate(sheet.iter_rows(), 1):
            html += "<tr>"
            for col_idx, cell in enumerate(row, 1):
                if (row_idx, col_idx) in merged_set:
                    continue
                val = str(cell.value) if cell.value is not None else ""
                rspan, cspan = merge_info.get((row_idx, col_idx), (1, 1))

                if row_idx == 1:
                    html += f'<th rowspan="{rspan}" colspan="{cspan}">{val}</th>'
                else:
                    html += f'<td rowspan="{rspan}" colspan="{cspan}">{val}</td>'
            html += "</tr>"

        html += """
            </table>
        </body>
        </html>
        """
        wb.close()
        return html

    # ------------------------------
    # 导入 HTML 文件 → 浏览器渲染
    # ------------------------------
    def import_html(self):
        path, _ = QFileDialog.getOpenFileName(filter="HTML (*.html)")
        if path:
            with open(path, "r", encoding="utf-8") as f:
                html = f.read()
            self.web.setHtml(html)
            QMessageBox.information(self, "成功", "HTML 已通过 Web 引擎渲染")

    # ------------------------------
    # 导出 HTML
    # ------------------------------
    def export_html(self):
        path, _ = QFileDialog.getSaveFileName(filter="HTML (*.html)")
        if path:
            html = self.web.page().toHtml(lambda html: self.save_file(html, path))

    def save_file(self, html, path):
        with open(path, "w", encoding="utf-8") as f:
            f.write(html)
        QMessageBox.information(self, "成功", "HTML 已导出")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    w = ExcelWebEngineViewer()
    w.show()
    sys.exit(app.exec_())